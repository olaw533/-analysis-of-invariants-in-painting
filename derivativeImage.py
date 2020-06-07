import sys
import cv2 as cv
import numpy as np
from openpyxl import Workbook
import pandas

def main(argv):
    workbook = Workbook('avgDEV.xlsx')

    faf(workbook.create_sheet('Beksinski'), 'Base\\Beksinski\\')
    faf(workbook.create_sheet('Hockney'), 'Base\\Hockney\\')
    faf(workbook.create_sheet('vanGogh'), 'Base\\vanGogh\\')
    faf(workbook.create_sheet('Modigliani'), 'Base\\Modigliani\\')
    faf(workbook.create_sheet('Monet'), 'Base\\Monet\\')

    workbook.save('avgDEV.xlsx')
    workbook.close()

    cv.waitKey(0)
    return 0

def faf(worksheet, path):
    ddepth = cv.CV_16S
    kernel_size = 3
    window_name = "Laplace Demo"

    worksheet.append(['Kolorowe', 'Wszystkie', 'Procenty'])

    for i in range(1, 51):
        imagePath = path + str(i) + '.jpg'
        src = cv.imread(cv.samples.findFile(imagePath), cv.IMREAD_COLOR)
        if src is None:
            print ('Error opening image')
            print ('Program Arguments: [image_name -- default lena.jpg]')
            return -1

        src = cv.GaussianBlur(src, (3, 3), 0)
        src_gray = cv.cvtColor(src, cv.COLOR_BGR2GRAY)
       # cv.namedWindow(window_name, cv.WINDOW_AUTOSIZE)
        dst = cv.Laplacian(src_gray, ddepth, ksize=kernel_size)
        abs_dst = cv.convertScaleAbs(dst)
        #cv.imshow(window_name, abs_dst)
        #cv.imwrite('edgeImage.jpg', abs_dst)
        #counting(imagePath)

        # img_name = 'edgeImage.jpg'

        boundaries = [
            (0, 255)
        ]

        for (lower, upper) in boundaries:
            lower = np.array(lower, dtype="uint8")
            upper = np.array(upper, dtype="uint8")
            mask = cv.inRange(abs_dst, lower, upper)
            output = cv.bitwise_and(abs_dst, abs_dst, mask=mask)

        tot_pixel = output.size
        red_pixel = np.count_nonzero(output)
        percentage = round(red_pixel * 100 / tot_pixel, 2)

        print("Color pixels: " + str(red_pixel))
        print("Total pixels: " + str(tot_pixel))
        print("Percentage of color pixels: " + str(percentage) + "%")

        worksheet.append([red_pixel, tot_pixel, percentage])


if __name__ == "__main__":
    main(sys.argv[1:])

